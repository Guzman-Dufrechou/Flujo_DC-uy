import numpy as np
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import xlsxwriter
import matplotlib.pyplot as plt
import pandas as pd
###############################################
###############################################
def guardar_PTDF(archivo_excel,nombre_pagina,modelo_ptdf,red):
    '''
    Construye la hoja SolucionTareas
    '''
    pagina = archivo_excel.create_sheet(title=nombre_pagina)
    nombre_barras = [red.barras['NAME'].tolist()]
    nombre_ramas = []
    
    for i in red.ramas.index:
        nombre=str(red.ramas['FROMNUMBER'][i]).strip()+' / '+str(red.ramas['TONUMBER'][i]).strip() + ' ' + '('+str(red.ramas['ID'][i]).strip()+')'
        #los parentesis rectos [] son para que se agregue una lista en el ultimo dato de la lista
        nombre_ramas.append([nombre])

    pagina.cell(row=3,column=1).value = 'Desde / Hasta (ID)'
    pagina.cell(row=1,column=3).value = 'Nombre barra'
    agregar_list(pagina,nombre_barras,1,2)
    agregar_list(pagina,nombre_ramas,2,1)
    agregar_list(pagina,modelo_ptdf.tolist(),2,2)

def guardar_Flujos(archivo_excel,nombre_pagina,matriz_de_flujos,red,lista_de_modelos):
    '''
    Construye la hoja SolucionTareas
    '''
    pagina = archivo_excel.create_sheet(title=nombre_pagina)
    
    nombre_ramas = []
    
    for i in red.ramas.index:
        nombre=str(red.ramas['FROMNAME'][i]).strip()+' / '+str(red.ramas['TONAME'][i]).strip() + ' ' + '('+str(red.ramas['ID'][i]).strip()+')'
        #los parentesis rectos [] son para que se agregue una lista en el ultimo dato de la lista
        nombre_ramas.append([nombre])
 
    pagina.cell(row=3,column=1).value = 'Desde / Hasta (ID)'
    pagina.cell(row=1,column=3).value = 'Modelos: '
    agregar_list(pagina,lista_de_modelos,1,2)
    agregar_list(pagina,nombre_ramas,2,1)
    agregar_list(pagina,matriz_de_flujos.tolist(),2,2)

###############################################

def De_R1C1_A_A1(i,j):
  '''
  Pasa de coordenadas (i,j) a las coordenadas que usa la hoja de datos Li (donde L es la letra equivalente a la j)
  '''
  col=xlsxwriter.utility.xl_col_to_name(j)
  string=col+str(i)
  return string
  
def colorear_grupo(lista,color_hex,fill_type):
  for rows in lista:
    for cell in rows:
        cell.fill = PatternFill(start_color=color_hex.color.rgb,end_color=color_hex.color.rgb, fill_type = fill_type)
  return
  
def color_style(r,g,b):
  color_r = f"{r:02X}{g:02X}{b:02X}"
  font = Font(color=color_r)
  return font 

def Borrar_hoja(Resultado, name_hoja):
  try:
    delete = Resultado.get_sheet_by_name(name_hoja)
    Resultado.remove_sheet(delete)
  except:
    pass
  return

def agregar_list(hoja,matriz,inicio_filas,inicio_columna):
  for i in range(len(matriz)):
    for j in range(len(matriz[i])):
       hoja.cell(row=inicio_filas+i+1,column=inicio_columna+j+1).value = matriz[i][j]

def graficar_resultados(matriz_de_flujos , red , lista_de_flujos,lista_de_errores):
  nombre_ramas = []
  #for i in red.ramas[:red.numero_ramas].index:
  for i in red.ramas.index:
      nombre=str(red.ramas['FROMNAME'][i]).strip()+' / '+str(red.ramas['TONAME'][i]).strip() + ' ' + '('+str(red.ramas['ID'][i]).strip()+')'   
      #los parentesis rectos [] son para que se agregue una lista en el ultimo dato de la lista
      nombre_ramas.append(nombre)
  datos = pd.DataFrame(data = matriz_de_flujos, columns = lista_de_flujos + lista_de_errores,index = nombre_ramas)

  fig = plt.figure('Comparacion de Flujos')
  plt.title('Figura 1: Comparación de flujos')
  plt.grid(color='lightgray',zorder=0)
  width = 1/(len(lista_de_flujos)+1)
  n = len(datos.index)
  x = np.arange(n)
  
  for i,modelo in enumerate(lista_de_flujos):
    plt.bar(x+i*width,datos[modelo], width = width, label = modelo,zorder=3)
    
  plt.xticks(x,datos.index,rotation = 'vertical')
  plt.tick_params(axis='x', labelsize=7)
  plt.ylabel('Potencia (MW)')
  plt.xlabel('Ramas')
  plt.legend(loc='best')

  fig = plt.figure('Comparacion de errores')
  plt.title('Figura 2: Comparación de errores respecto al flujo AC')
  plt.grid(color='lightgray',zorder=0)
  width = 1/(len(lista_de_errores))
  n = len(datos.index)
  x = np.arange(n)

  for i,modelo in enumerate(lista_de_errores[1:]):
    plt.bar(x+i*width,datos[modelo], width = width, label = modelo,zorder=3)

  plt.xticks(x,datos.index,rotation = 'vertical')
  plt.tick_params(axis='x', labelsize=7)
  plt.legend(loc='best')
  plt.ylabel('Error (MW)')
  plt.xlabel('Ramas')
  plt.show()
